from openpyxl import Workbook
import random

excel_book = Workbook()
excel_name = "Эксель.xlsx"

excel_sheet = excel_book.create_sheet(title="Рандом")

for i in range(1, 13):
    for j in range(1, 13):
        excel_sheet.cell(column=i, row=j, value=random.randint(1, 100))

excel_book.save(filename=excel_name)