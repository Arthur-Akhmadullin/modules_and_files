import requests
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook, load_workbook

#Первое задание
response = requests.post('http://httpbin.org/post', data = {'UserId':'12345', 'Status':'On'})
if response.status_code == 200:
    print("Данные отправлены корректно")
else:
    print("Ошибка отправки данных", response.status_code)

#Второе задание
response = requests.get('https://www.chitai-gorod.ru/catalog/books/9170')
if response.status_code == 200:
    book = BeautifulSoup(response.text)
else:
    print("Ошибка", response.status_code)

#Результаты сохраняем в списки
bookname_list = []
author_list = []
price_list = []

for items in book.find_all('div', {'class': 'product-card__title js-analytic-product-title'}):
    bookname_list.append(items.text.strip())
print(bookname_list)

for items in book.find_all('div', {'class': 'product-card__author'}):
    author_list.append(items.text.strip())
print(author_list)

for items in book.find_all('span', {'class': 'product-price__value'}):
    price_list.append(items.text)
print(price_list)

#Объединияем в словарь
book_dict = {}
for i in range(len(price_list)):
    book_dict["id"+str(i+1)] = {
        "author": author_list[i],
        "bookname": bookname_list[i],
        "price": price_list[i]
    }

#Словарь сохраняем в json-файл
with open("file.json", "w", encoding="utf-8") as f:
    json.dump(book_dict, f, indent=2, ensure_ascii=False)

#А также можно сохранить в формате .xslx
wb = Workbook()
dest_filename = 'books.xlsx'
book_sheet = wb.create_sheet(title="books")

book_sheet.cell(column=1, row=1, value="Identificator")
book_sheet.cell(column=2, row=1, value="Author")
book_sheet.cell(column=3, row=1, value="Bookname")
book_sheet.cell(column=4, row=1, value="Price")

for i in range(len(price_list)):
    book_sheet.cell(column=1, row=i+2, value="id"+str(i+1))
    book_sheet.cell(column=2, row=i+2, value=author_list[i])
    book_sheet.cell(column=3, row=i+2, value=bookname_list[i])
    book_sheet.cell(column=4, row=i+2, value=price_list[i])

wb.save(filename=dest_filename)